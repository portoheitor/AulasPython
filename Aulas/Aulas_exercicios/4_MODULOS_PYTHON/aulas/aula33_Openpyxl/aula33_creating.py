"""
# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORDBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = Workbook()
worksheet: Worksheet = workbook.active

#criando Cabeçalhos
worksheet.cell(1, 1, "Nome")
worksheet.cell(1, 2, "Idade")
worksheet.cell(1, 3, "Nota")

students = [
    ['Heitor', 30, 8.8],
    ['Maite', 1, 10.0],
    ['Arielle', 29, 6.8],
    ['jose', 54, 7.8],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_collum in enumerate(student_row, start=1):
#         print(i, j, student_collum)

for student in students:
    worksheet.append(student)

workbook.save(WORDBOOK_PATH)
